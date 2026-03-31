"""导入相关路由。

这里实际上维护两条导入链路：

1. Excel 导入
   - 下载模板
   - 上传项目/合同/付款三类 Excel
   - 按行做校验并落库
2. 截图识别导入
   - 上传多张 OA / 合同截图给 AI 识别
   - 前端人工确认识别结果
   - 再由确认接口统一写入数据库
"""

from __future__ import annotations

import base64
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from .. import models, schemas
from ..database import get_db
from ..services.ai_parser import AIParserError, parse_screenshots

router = APIRouter(prefix="/api/import", tags=["导入"])

# Excel 模板、表头映射、示例行和必填字段都集中在这里维护。
# 这样模板下载和实际导入解析可以共用同一份配置，避免两边字段漂移。
ENTITY_CONFIG: dict[str, dict[str, Any]] = {
    "projects": {
        "headers": [
            (["项目编号"], "project_code"),
            (["项目名称"], "project_name"),
            (["项目属性", "项目类型"], "project_type"),
            (["立项日期", "申请日期"], "start_date"),
            (["项目状态", "是否作废"], "status"),
            (["项目金额", "预计项目金额"], "budget"),
            (["负责人", "项目负责人"], "manager"),
            (["备注"], "remark"),
        ],
        "example": {
            "project_code": "WLGS202400076",
            "project_name": "示例项目",
            "project_type": "研发项目",
            "start_date": "2026-03-01",
            "status": "立项",
            "budget": "1000000",
            "manager": "张三",
            "remark": "示例备注",
        },
        "required": ["project_code", "project_name"],
    },
    "contracts": {
        "headers": [
            (["项目编号"], "project_code"),
            (["合同编号"], "contract_code"),
            (["合同名称"], "contract_name"),
            (["采购类型"], "procurement_type"),
            (["费用归属责任中心"], "cost_department"),
            (["供应商"], "vendor"),
            (["合同金额"], "amount"),
            (["合同金额（变更前）"], "amount_before_change"),
            (["签订日期"], "sign_date"),
            (["备案日期"], "filing_date"),
            (["开始执行日期"], "start_date"),
            (["结束执行日期"], "end_date"),
            (["主合同编号"], "parent_contract_code"),
            (["合同续签类型"], "renewal_type"),
            (["收支方向"], "payment_direction"),
            (["合同状态"], "status"),
            (["备案文件"], "filing_reference"),
            (["备注"], "remark"),
        ],
        "example": {
            "project_code": "WLGS202400076",
            "contract_code": "WLGS202500056CG20260005",
            "contract_name": "示例合同",
            "procurement_type": "项目类",
            "cost_department": "研发部",
            "vendor": "示例供应商",
            "amount": "250000",
            "amount_before_change": "200000",
            "sign_date": "2026-03-10",
            "filing_date": "2026-03-12",
            "start_date": "2026-03-15",
            "end_date": "2026-12-31",
            "parent_contract_code": "",
            "renewal_type": "固定期限合同",
            "payment_direction": "支出",
            "status": "签订",
            "filing_reference": "备案文件001",
            "remark": "示例备注",
        },
        "required": ["project_code", "contract_code", "contract_name", "amount", "status"],
    },
    "payments": {
        "headers": [
            (["合同编号"], "contract_code"),
            (["序号"], "seq"),
            (["付款阶段"], "phase"),
            (["计划日期"], "planned_date"),
            (["计划金额"], "planned_amount"),
            (["实际日期"], "actual_date"),
            (["实际金额"], "actual_amount"),
            (["付款状态"], "payment_status"),
            (["支付说明"], "description"),
            (["备注"], "remark"),
        ],
        "example": {
            "contract_code": "WLGS202500056CG20260005",
            "seq": "1",
            "phase": "第一期",
            "planned_date": "2026-04-01",
            "planned_amount": "80000",
            "actual_date": "",
            "actual_amount": "",
            "payment_status": "未付",
            "description": "首付款",
            "remark": "示例备注",
        },
        "required": ["contract_code", "payment_status"],
    },
}


# 下面几个工具函数负责把 Excel / AI 返回的"弱类型文本值"
# 统一转成数据库层需要的 Decimal / date / 业务字段。
def _to_decimal(value: Any) -> Decimal | None:
    """将输入值转换为 Decimal。"""
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation as exc:
        raise HTTPException(status_code=400, detail="金额字段格式不正确") from exc


def _build_pending_amount(planned_amount: Decimal | None, actual_amount: Decimal | None) -> Decimal | None:
    """计算待付款金额。"""
    if planned_amount is None and actual_amount is None:
        return None
    return (planned_amount or Decimal("0")) - (actual_amount or Decimal("0"))


def _to_date(value: Any) -> date | None:
    """将输入值转换为日期，兼容多种常见格式。"""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise HTTPException(status_code=400, detail=f"无法识别日期格式：{s}，建议使用 YYYY-MM-DD")


def _normalize_project_info(contract_data: dict[str, Any]) -> tuple[str, str]:
    """规范化项目编号和项目名称。"""
    project_code = (contract_data.get("project_code") or "").strip()
    project_name = (contract_data.get("project_name") or "").strip()

    if not project_code and not project_name:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        project_code = f"AI-{timestamp}"
        project_name = f"AI导入项目-{timestamp}"
    elif not project_code:
        project_code = f"AI-{datetime.now().strftime('%Y%m%d%H%M%S')}"
    elif not project_name:
        project_name = project_code

    return project_code, project_name


def _get_or_create_project(contract_data: dict[str, Any], db: Session) -> models.Project:
    """通过项目编号查找项目，不存在时自动创建。"""
    project_code, project_name = _normalize_project_info(contract_data)

    project = db.query(models.Project).filter(models.Project.project_code == project_code).first()
    if project:
        # AI 导入场景下，可能第一次只识别出了项目编号，后续补录时再把名称补齐。
        if not project.project_name and project_name:
            project.project_name = project_name
        return project

    project = models.Project(
        project_code=project_code,
        project_name=project_name,
        status="立项",
    )
    db.add(project)
    db.flush()
    return project


def _ensure_entity(entity: str) -> dict[str, Any]:
    """校验导入实体类型。"""
    config = ENTITY_CONFIG.get(entity)
    if not config:
        raise HTTPException(status_code=400, detail="entity 仅支持 projects/contracts/payments")
    return config


@router.get("/template/{entity}")
def download_template(entity: str):
    """下载 Excel 导入模板。"""
    config = _ensure_entity(entity)

    # 模板只放两行：表头 + 示例，尽量保持轻量，方便业务人员直接另存后填写。
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "导入模板"
    worksheet.append([titles[0] for titles, _ in config["headers"]])
    worksheet.append([config["example"].get(field, "") for _, field in config["headers"]])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{entity}_template.xlsx"'},
    )


@router.post("/excel/{entity}")
async def import_excel(
    entity: str,
    file: UploadFile = File(...),
    duplicate_action: str = Query(default="skip", pattern="^(skip|update)$"),
    db: Session = Depends(get_db),
):
    """导入 Excel 数据。"""
    config = _ensure_entity(entity)

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持上传 xlsx 文件")

    try:
        content = await file.read()
        workbook = load_workbook(BytesIO(content))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Excel 文件解析失败：{exc}") from exc

    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="模板至少需要表头和一行数据")

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    header_index = {header: index for index, header in enumerate(headers)}

    # 每个字段取第一个能匹配上的列名；必填字段匹配不上才报错，可选字段缺失则置 None。
    field_col: dict[str, int] = {}
    required_fields = set(config["required"])
    missing_required: list[str] = []
    for titles, field in config["headers"]:
        matched = next((t for t in titles if t in header_index), None)
        if matched:
            field_col[field] = header_index[matched]
        elif field in required_fields:
            missing_required.append(titles[0])
    if missing_required:
        raise HTTPException(status_code=400, detail=f"缺少必填列：{', '.join(missing_required)}")

    # 结果按"成功 / 失败 / 跳过"累计，前端可以直接把它展示成导入报告。
    result = {"success": 0, "failed": 0, "skipped": 0, "errors": []}

    for row_number, row in enumerate(rows[1:], start=2):
        row_data: dict[str, Any] = {}
        for _, field in config["headers"]:
            idx = field_col.get(field)
            row_data[field] = row[idx] if idx is not None else None

        # "是否作废"列的值兼容：是→作废，否→立项，其余原样保留。
        if "status" in row_data and row_data["status"] in ("是", "否"):
            row_data["status"] = "作废" if row_data["status"] == "是" else "立项"

        if all(value in (None, "") for value in row_data.values()):
            continue

        try:
            for required_field in config["required"]:
                if row_data.get(required_field) in (None, ""):
                    raise ValueError(f"{required_field} 不能为空")

            if entity == "projects":
                # 项目以 project_code 为幂等键。
                project_code = str(row_data["project_code"]).strip()
                existing = db.query(models.Project).filter(models.Project.project_code == project_code).first()
                payload = {
                    "project_code": project_code,
                    "project_name": str(row_data["project_name"]).strip(),
                    "project_type": row_data.get("project_type"),
                    "start_date": _to_date(row_data.get("start_date")),
                    "status": str(row_data["status"]).strip() if row_data.get("status") else "立项",
                    "budget": _to_decimal(row_data.get("budget")),
                    "manager": row_data.get("manager"),
                    "remark": row_data.get("remark"),
                }
                if existing:
                    if duplicate_action == "skip":
                        result["skipped"] += 1
                        continue
                    for key, value in payload.items():
                        setattr(existing, key, value)
                else:
                    db.add(models.Project(**payload))

            elif entity == "contracts":
                # 合同必须先能找到所属项目，不能脱离项目独立导入。
                project_code = str(row_data["project_code"]).strip()
                project = db.query(models.Project).filter(models.Project.project_code == project_code).first()
                if not project:
                    raise ValueError("项目编号不存在")

                # 合同以 contract_code 为幂等键。
                contract_code = str(row_data["contract_code"]).strip()
                existing = db.query(models.Contract).filter(models.Contract.contract_code == contract_code).first()
                payload = {
                    "project_id": project.id,
                    "contract_code": contract_code,
                    "contract_name": str(row_data["contract_name"]).strip(),
                    "procurement_type": row_data.get("procurement_type"),
                    "cost_department": row_data.get("cost_department"),
                    "vendor": row_data.get("vendor"),
                    "amount": _to_decimal(row_data.get("amount")),
                    "amount_before_change": _to_decimal(row_data.get("amount_before_change")),
                    "sign_date": _to_date(row_data.get("sign_date")),
                    "filing_date": _to_date(row_data.get("filing_date")),
                    "start_date": _to_date(row_data.get("start_date")),
                    "end_date": _to_date(row_data.get("end_date")),
                    "parent_contract_code": row_data.get("parent_contract_code"),
                    "renewal_type": row_data.get("renewal_type"),
                    "payment_direction": row_data.get("payment_direction"),
                    "status": str(row_data["status"]).strip(),
                    "filing_reference": row_data.get("filing_reference"),
                    "remark": row_data.get("remark"),
                }
                if existing:
                    if duplicate_action == "skip":
                        result["skipped"] += 1
                        continue
                    for key, value in payload.items():
                        setattr(existing, key, value)
                else:
                    db.add(models.Contract(**payload))

            else:
                # 付款以"合同 + 序号"识别重复记录；没有序号时只能新增。
                contract_code = str(row_data["contract_code"]).strip()
                contract = db.query(models.Contract).filter(models.Contract.contract_code == contract_code).first()
                if not contract:
                    raise ValueError("合同编号不存在")

                seq = int(row_data["seq"]) if row_data.get("seq") not in (None, "") else None
                existing = None
                if seq is not None:
                    existing = (
                        db.query(models.Payment)
                        .filter(models.Payment.contract_id == contract.id, models.Payment.seq == seq)
                        .first()
                    )

                planned_amount = _to_decimal(row_data.get("planned_amount"))
                actual_amount = _to_decimal(row_data.get("actual_amount"))
                payload = {
                    "contract_id": contract.id,
                    "seq": seq,
                    "phase": row_data.get("phase"),
                    "planned_date": _to_date(row_data.get("planned_date")),
                    "planned_amount": planned_amount,
                    "actual_date": _to_date(row_data.get("actual_date")),
                    "actual_amount": actual_amount,
                    "pending_amount": _build_pending_amount(planned_amount, actual_amount),
                    "payment_status": str(row_data["payment_status"]).strip(),
                    "description": row_data.get("description"),
                    "remark": row_data.get("remark"),
                }
                if existing:
                    if duplicate_action == "skip":
                        result["skipped"] += 1
                        continue
                    for key, value in payload.items():
                        setattr(existing, key, value)
                else:
                    db.add(models.Payment(**payload))

            # 每一行单独提交，保证部分成功时不会因为一行脏数据拖垮整个文件。
            db.commit()
            result["success"] += 1
        except Exception as exc:  # noqa: BLE001
            # 保留行号和错误文本，方便前端把失败明细直接反馈给用户。
            db.rollback()
            result["failed"] += 1
            result["errors"].append({"row": row_number, "message": str(exc)})

    return result


@router.post("/screenshot", response_model=schemas.AIScreenshotParseResponse)
async def import_from_screenshot(files: list[UploadFile] = File(...)):
    """上传多张截图并调用 Claude Vision 解析。"""
    if not files:
        raise HTTPException(status_code=400, detail="请至少上传一张截图")

    # AI 服务接收的是图片内容，这里统一转成 base64 文本数组后再交给解析器。
    images_b64: list[str] = []
    for file in files:
        content = await file.read()
        if content:
            images_b64.append(base64.b64encode(content).decode("utf-8"))

    if not images_b64:
        raise HTTPException(status_code=400, detail="上传的截图内容为空")

    try:
        # 这里只做识别，不做数据库写入。真正落库放在 confirm 接口，避免 AI 误识别直接污染数据。
        parsed_data, uncertain_fields = parse_screenshots(images_b64)
    except AIParserError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"AI 识别失败：{exc}") from exc

    return schemas.AIScreenshotParseResponse(parsed_data=parsed_data, uncertain_fields=uncertain_fields)


@router.post("/screenshot/confirm")
def confirm_screenshot_import(payload: schemas.AIScreenshotConfirmRequest, db: Session = Depends(get_db)):
    """确认识别结果并写入数据库。"""
    parsed_data = payload.parsed_data
    contract_data = parsed_data.get("contract") or {}
    items = parsed_data.get("items") or []
    payment_plans = parsed_data.get("payment_plans") or []
    changes = parsed_data.get("changes") or []

    contract_code = (contract_data.get("contract_code") or "").strip()
    contract_name = (contract_data.get("contract_name") or "").strip()
    status = (contract_data.get("status") or "").strip()

    if not contract_code:
        raise HTTPException(status_code=400, detail="合同编号不能为空")
    if not contract_name:
        raise HTTPException(status_code=400, detail="合同名称不能为空")
    if not status:
        raise HTTPException(status_code=400, detail="合同状态不能为空")
    if _to_decimal(contract_data.get("amount")) is None:
        raise HTTPException(status_code=400, detail="合同金额不能为空")

    # 截图识别导入采用"两段式提交"：
    # 1. screenshot 接口只返回识别结果
    # 2. confirm 接口在人工确认后一次性创建项目 / 合同 / 子表
    existing_contract = db.query(models.Contract).filter(models.Contract.contract_code == contract_code).first()
    if existing_contract:
        raise HTTPException(status_code=400, detail="合同编号已存在，请修改后再导入")

    # 如果截图里没有完整项目信息，会自动兜底创建一个 AI 导入项目，避免合同成为孤儿数据。
    project = _get_or_create_project(contract_data, db)

    contract = models.Contract(
        project_id=project.id,
        contract_code=contract_code,
        contract_name=contract_name,
        procurement_type=contract_data.get("procurement_type"),
        cost_department=contract_data.get("cost_department"),
        vendor=contract_data.get("vendor"),
        amount=_to_decimal(contract_data.get("amount")),
        amount_before_change=_to_decimal(contract_data.get("amount_before_change")),
        sign_date=_to_date(contract_data.get("sign_date")),
        filing_date=_to_date(contract_data.get("filing_date")),
        start_date=_to_date(contract_data.get("start_date")),
        end_date=_to_date(contract_data.get("end_date")),
        parent_contract_code=contract_data.get("parent_contract_code"),
        renewal_type=contract_data.get("renewal_type"),
        payment_direction=contract_data.get("payment_direction"),
        status=status,
        filing_reference=contract_data.get("filing_reference"),
        remark=contract_data.get("remark"),
    )
    db.add(contract)
    db.flush()

    # 标的、付款计划、变更记录都挂在新合同下面一起落库，保证导入后的详情页数据是完整的。
    for index, item in enumerate(items, start=1):
        db.add(
            models.ContractItem(
                contract_id=contract.id,
                seq=int(item.get("seq") or index),
                item_name=item.get("item_name") or f"标的{index}",
                quantity=_to_decimal(item.get("quantity")) or Decimal("0"),
                unit=item.get("unit"),
                unit_price=_to_decimal(item.get("unit_price")),
                amount=_to_decimal(item.get("amount")) or Decimal("0"),
            )
        )

    for index, payment in enumerate(payment_plans, start=1):
        planned_amount = _to_decimal(payment.get("planned_amount"))
        actual_amount = _to_decimal(payment.get("actual_amount"))
        db.add(
            models.Payment(
                contract_id=contract.id,
                seq=int(payment["seq"]) if payment.get("seq") not in (None, "") else index,
                phase=payment.get("phase"),
                planned_date=_to_date(payment.get("planned_date")),
                planned_amount=planned_amount,
                actual_date=_to_date(payment.get("actual_date")),
                actual_amount=actual_amount,
                pending_amount=_build_pending_amount(planned_amount, actual_amount),
                payment_status=payment.get("payment_status") or "未付",
                description=payment.get("description"),
                remark=payment.get("remark"),
            )
        )

    for index, change in enumerate(changes, start=1):
        change_date = _to_date(change.get("change_date"))
        # 没有变更日期的记录通常不能成立，这里直接跳过而不是写入脏数据。
        if not change_date:
            continue
        db.add(
            models.ContractChange(
                contract_id=contract.id,
                seq=int(change.get("seq") or index),
                change_date=change_date,
                change_info=change.get("change_info"),
                before_content=change.get("before_content"),
                after_content=change.get("after_content"),
                change_description=change.get("change_description"),
            )
        )

    db.commit()
    db.refresh(contract)

    return {
        "message": "导入成功",
        "project_id": project.id,
        "contract_id": contract.id,
    }
