from __future__ import annotations

from datetime import date, timedelta
from decimal import Decimal
from io import BytesIO
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook


def unique_code(prefix: str) -> str:
    return f"{prefix}-{uuid4().hex[:8]}"


def as_decimal(value: object) -> Decimal:
    return Decimal(str(value))


def build_excel(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class ApiFactory:
    def __init__(self, client: TestClient):
        self.client = client

    def create_project(self, **overrides: object) -> dict[str, object]:
        payload = {
            "project_code": unique_code("PJ"),
            "project_name": unique_code("项目"),
            "project_type": "研发项目",
            "start_date": date.today().isoformat(),
            "status": "立项",
            "budget": "100000.00",
            "manager": "测试负责人",
            "remark": "pytest smoke",
        }
        payload.update(overrides)
        response = self.client.post("/api/projects", json=payload)
        assert response.status_code == 200, response.text
        return response.json()

    def create_contract(self, project_id: int, **overrides: object) -> dict[str, object]:
        payload = {
            "project_id": project_id,
            "contract_code": unique_code("HT"),
            "contract_name": unique_code("测试合同"),
            "vendor": "测试供应商",
            "amount": "50000.00",
            "status": "签订",
            "payment_direction": "支出",
            "items": [],
            "payments": [],
        }
        payload.update(overrides)
        response = self.client.post("/api/contracts", json=payload)
        assert response.status_code == 200, response.text
        return response.json()

    def create_payment(self, contract_id: int, **overrides: object) -> dict[str, object]:
        payload = {
            "contract_id": contract_id,
            "seq": 1,
            "phase": "首付款",
            "planned_date": (date.today() + timedelta(days=3)).isoformat(),
            "planned_amount": "12000.00",
            "actual_amount": "0",
            "payment_status": "未付",
            "description": "pytest smoke",
            "remark": "自动创建",
        }
        payload.update(overrides)
        response = self.client.post("/api/payments", json=payload)
        assert response.status_code == 200, response.text
        return response.json()


@pytest.fixture()
def factory(client: TestClient) -> ApiFactory:
    return ApiFactory(client)


def test_health_and_dashboard_endpoints(client: TestClient, factory: ApiFactory) -> None:
    project = factory.create_project()
    contract = factory.create_contract(project["id"])
    payment = factory.create_payment(contract["id"])

    health_response = client.get("/")
    assert health_response.status_code == 200
    assert health_response.json() == {"message": "后端服务运行正常"}

    summary_response = client.get("/api/dashboard/summary")
    assert summary_response.status_code == 200
    summary = summary_response.json()
    assert summary["project_count"] >= 1
    assert summary["contract_count"] >= 1
    assert summary["payment_count"] >= 1
    assert "project_status_distribution" in summary

    workflow_response = client.get("/api/dashboard/workflow")
    assert workflow_response.status_code == 200
    workflow = workflow_response.json()
    assert workflow["project_stage"]["total"] >= 1
    assert workflow["contract_stage"]["total"] >= 1
    assert workflow["payment_stage"]["total"] >= 1

    analysis_response = client.get("/api/dashboard/analysis")
    assert analysis_response.status_code == 200
    analysis = analysis_response.json()
    assert analysis["financial_health"]["total_contract_amount"] >= 50000
    assert analysis["coverage"]["project_contract_link_rate"] > 0
    assert analysis["payment_risk"]["due_soon_count"] >= 1
    assert len(analysis["top_risk_projects"]) >= 1
    assert len(analysis["priority_payments"]) >= 1

    pending_response = client.get("/api/dashboard/pending-payments")
    assert pending_response.status_code == 200
    pending_rows = pending_response.json()
    assert any(row["id"] == payment["id"] for row in pending_rows)


def test_project_crud_and_delete_protection(client: TestClient, factory: ApiFactory) -> None:
    standalone_project = factory.create_project()

    get_response = client.get(f"/api/projects/{standalone_project['id']}")
    assert get_response.status_code == 200
    assert get_response.json()["project_code"] == standalone_project["project_code"]

    list_response = client.get(
        "/api/projects",
        params={"page": 1, "page_size": 10, "search": standalone_project["project_code"]},
    )
    assert list_response.status_code == 200
    list_data = list_response.json()
    assert list_data["total"] >= 1
    assert any(item["id"] == standalone_project["id"] for item in list_data["items"])

    update_response = client.put(
        f"/api/projects/{standalone_project['id']}",
        json={"status": "合同", "manager": "已更新负责人"},
    )
    assert update_response.status_code == 200
    updated_project = update_response.json()
    assert updated_project["status"] == "合同"
    assert updated_project["manager"] == "已更新负责人"

    delete_response = client.delete(f"/api/projects/{standalone_project['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["message"] == "删除成功"
    assert client.get(f"/api/projects/{standalone_project['id']}").status_code == 404

    protected_project = factory.create_project()
    factory.create_contract(protected_project["id"])

    protected_delete_response = client.delete(f"/api/projects/{protected_project['id']}")
    assert protected_delete_response.status_code == 400
    assert protected_delete_response.json()["message"] == "项目下存在合同，禁止删除"


def test_contract_crud_subresources_and_warnings(client: TestClient, factory: ApiFactory) -> None:
    project = factory.create_project()
    contract = factory.create_contract(project["id"], amount="1000.00")
    contract_id = int(contract["id"])

    item_response = client.post(
        f"/api/contracts/{contract_id}/items",
        json={
            "seq": 1,
            "item_name": "服务器",
            "quantity": "2",
            "unit": "台",
            "unit_price": "300.00",
            "amount": "600.00",
        },
    )
    assert item_response.status_code == 200
    item = item_response.json()

    update_item_response = client.put(
        f"/api/contracts/{contract_id}/items/{item['id']}",
        json={"amount": "900.00", "unit_price": "450.00"},
    )
    assert update_item_response.status_code == 200
    assert as_decimal(update_item_response.json()["amount"]) == Decimal("900.00")

    change_response = client.post(
        f"/api/contracts/{contract_id}/changes",
        json={
            "seq": 1,
            "change_date": date.today().isoformat(),
            "change_info": "范围扩大",
            "before_content": "基础服务",
            "after_content": "基础服务+运维",
            "change_description": "新增运维内容",
        },
    )
    assert change_response.status_code == 200
    change = change_response.json()

    update_change_response = client.put(
        f"/api/contracts/{contract_id}/changes/{change['id']}",
        json={"change_info": "范围二次扩大"},
    )
    assert update_change_response.status_code == 200
    assert update_change_response.json()["change_info"] == "范围二次扩大"

    payment = factory.create_payment(contract_id, planned_amount="1000.00", actual_amount="1200.00", payment_status="已付款")

    contract_detail_response = client.get(f"/api/contracts/{contract_id}")
    assert contract_detail_response.status_code == 200
    contract_detail = contract_detail_response.json()
    assert "合同金额与标的清单合计不一致" in contract_detail["warnings"]
    assert "付款总额已超过合同金额" in contract_detail["warnings"]
    assert len(contract_detail["items"]) == 1
    assert len(contract_detail["payments"]) == 1
    assert len(contract_detail["changes"]) == 1

    delete_contract_response = client.delete(f"/api/contracts/{contract_id}")
    assert delete_contract_response.status_code == 400
    assert delete_contract_response.json()["message"] == "合同下存在子记录，禁止删除"

    assert client.delete(f"/api/contracts/{contract_id}/items/{item['id']}").status_code == 200
    assert client.delete(f"/api/contracts/{contract_id}/changes/{change['id']}").status_code == 200
    assert client.delete(f"/api/payments/{payment['id']}").status_code == 200

    final_delete_response = client.delete(f"/api/contracts/{contract_id}")
    assert final_delete_response.status_code == 200
    assert final_delete_response.json()["message"] == "删除成功"


def test_payment_crud_and_filtering(client: TestClient, factory: ApiFactory) -> None:
    project = factory.create_project()
    contract = factory.create_contract(project["id"])
    payment = factory.create_payment(contract["id"], planned_amount="1000.00", actual_amount="200.00", payment_status="已提报")

    get_response = client.get(f"/api/payments/{payment['id']}")
    assert get_response.status_code == 200
    assert get_response.json()["contract_id"] == contract["id"]

    filtered_response = client.get("/api/payments", params={"contract_id": contract["id"]})
    assert filtered_response.status_code == 200
    filtered_data = filtered_response.json()
    assert any(item["id"] == payment["id"] for item in filtered_data)

    update_response = client.put(
        f"/api/payments/{payment['id']}",
        json={
            "actual_amount": "600.00",
            "payment_status": "已付款",
            "actual_date": date.today().isoformat(),
        },
    )
    assert update_response.status_code == 200
    updated_payment = update_response.json()
    assert updated_payment["payment_status"] == "已付款"
    assert as_decimal(updated_payment["pending_amount"]) == Decimal("400.00")

    delete_response = client.delete(f"/api/payments/{payment['id']}")
    assert delete_response.status_code == 200
    assert delete_response.json()["message"] == "删除成功"
    assert client.get(f"/api/payments/{payment['id']}").status_code == 404


@pytest.mark.parametrize(
    ("entity", "headers"),
    [
        ("projects", ["项目编号", "项目名称", "项目属性", "状态", "预算", "负责人", "立项时间", "备注"]),
        ("contracts", ["合同编号", "合同名称", "项目编号", "项目名称", "供应商", "合同金额", "合同状态", "签订日期"]),
        ("payments", ["合同编号", "付款序号", "付款阶段", "计划日期", "计划金额", "实付日期", "实付金额", "付款状态", "备注"]),
    ],
)
def test_export_endpoints_return_xlsx(
    client: TestClient,
    factory: ApiFactory,
    entity: str,
    headers: list[str],
) -> None:
    project = factory.create_project()
    contract = factory.create_contract(project["id"])
    factory.create_payment(contract["id"])

    params: dict[str, object] = {"format": "xlsx"}
    if entity == "contracts":
        params["project_id"] = project["id"]
    elif entity == "payments":
        params["contract_id"] = contract["id"]

    response = client.get(f"/api/export/{entity}", params=params)
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    exported_headers = [cell.value for cell in worksheet[1]]
    assert exported_headers == headers
    assert worksheet.max_row >= 2


@pytest.mark.parametrize(
    ("entity", "expected_headers"),
    [
        ("projects", ["项目编号", "项目名称", "项目属性", "立项日期", "项目状态", "项目金额", "负责人", "备注"]),
        (
            "contracts",
            [
                "项目编号",
                "合同编号",
                "合同名称",
                "采购类型",
                "费用归属责任中心",
                "供应商",
                "合同金额",
                "合同金额（变更前）",
                "签订日期",
                "备案日期",
                "开始执行日期",
                "结束执行日期",
                "主合同编号",
                "合同续签类型",
                "收支方向",
                "合同状态",
                "备案文件",
                "备注",
            ],
        ),
        ("payments", ["合同编号", "序号", "付款阶段", "计划日期", "计划金额", "实际日期", "实际金额", "付款状态", "支付说明", "备注"]),
    ],
)
def test_import_templates_download(
    client: TestClient,
    entity: str,
    expected_headers: list[str],
) -> None:
    response = client.get(f"/api/import/template/{entity}")
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    worksheet = workbook.active
    actual_headers = [cell.value for cell in worksheet[1]]
    assert actual_headers == expected_headers
    assert worksheet.max_row >= 2


def test_excel_imports_support_skip_and_update(client: TestClient) -> None:
    project_code = unique_code("IMP-PJ")
    project_headers = ["项目编号", "项目名称", "项目属性", "立项日期", "项目状态", "项目金额", "负责人", "备注"]
    project_rows = [[project_code, "导入项目", "研发项目", date.today().isoformat(), "立项", "30000", "导入负责人", "首次导入"]]
    project_file = build_excel(project_headers, project_rows)

    project_import = client.post(
        "/api/import/excel/projects",
        params={"duplicate_action": "skip"},
        files={"file": ("projects.xlsx", project_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert project_import.status_code == 200
    assert project_import.json()["success"] == 1

    project_skip = client.post(
        "/api/import/excel/projects",
        params={"duplicate_action": "skip"},
        files={"file": ("projects.xlsx", project_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert project_skip.status_code == 200
    assert project_skip.json()["skipped"] == 1

    updated_project_file = build_excel(
        project_headers,
        [[project_code, "导入项目-更新", "服务项目", date.today().isoformat(), "合同", "45000", "更新负责人", "更新导入"]],
    )
    project_update = client.post(
        "/api/import/excel/projects",
        params={"duplicate_action": "update"},
        files={
            "file": (
                "projects-update.xlsx",
                updated_project_file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert project_update.status_code == 200
    assert project_update.json()["success"] == 1
    project_lookup = client.get("/api/projects", params={"page": 1, "page_size": 10, "search": project_code})
    project_item = next(item for item in project_lookup.json()["items"] if item["project_code"] == project_code)
    assert project_item["project_name"] == "导入项目-更新"
    assert project_item["manager"] == "更新负责人"

    contract_code = unique_code("IMP-HT")
    contract_headers = [
        "项目编号",
        "合同编号",
        "合同名称",
        "采购类型",
        "费用归属责任中心",
        "供应商",
        "合同金额",
        "合同金额（变更前）",
        "签订日期",
        "备案日期",
        "开始执行日期",
        "结束执行日期",
        "主合同编号",
        "合同续签类型",
        "收支方向",
        "合同状态",
        "备案文件",
        "备注",
    ]
    contract_file = build_excel(
        contract_headers,
        [
            [
                project_code,
                contract_code,
                "导入合同",
                "项目类",
                "研发中心",
                "原供应商",
                "18000",
                "",
                date.today().isoformat(),
                "",
                date.today().isoformat(),
                "",
                "",
                "固定期限合同",
                "支出",
                "签订",
                "备案-001",
                "首次导入",
            ]
        ],
    )
    contract_import = client.post(
        "/api/import/excel/contracts",
        params={"duplicate_action": "skip"},
        files={"file": ("contracts.xlsx", contract_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert contract_import.status_code == 200
    assert contract_import.json()["success"] == 1

    contract_update_file = build_excel(
        contract_headers,
        [
            [
                project_code,
                contract_code,
                "导入合同-更新",
                "项目类",
                "研发中心",
                "新供应商",
                "20000",
                "",
                date.today().isoformat(),
                "",
                date.today().isoformat(),
                "",
                "",
                "固定期限合同",
                "支出",
                "执行",
                "备案-002",
                "更新导入",
            ]
        ],
    )
    contract_update = client.post(
        "/api/import/excel/contracts",
        params={"duplicate_action": "update"},
        files={
            "file": (
                "contracts-update.xlsx",
                contract_update_file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert contract_update.status_code == 200
    assert contract_update.json()["success"] == 1
    contract_lookup = client.get("/api/contracts")
    imported_contract = next(item for item in contract_lookup.json() if item["contract_code"] == contract_code)
    assert imported_contract["contract_name"] == "导入合同-更新"
    assert imported_contract["vendor"] == "新供应商"

    payment_headers = ["合同编号", "序号", "付款阶段", "计划日期", "计划金额", "实际日期", "实际金额", "付款状态", "支付说明", "备注"]
    payment_file = build_excel(
        payment_headers,
        [[contract_code, 1, "第一期", (date.today() + timedelta(days=7)).isoformat(), "5000", "", "", "未付", "首付款", "首次导入"]],
    )
    payment_import = client.post(
        "/api/import/excel/payments",
        params={"duplicate_action": "skip"},
        files={"file": ("payments.xlsx", payment_file, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert payment_import.status_code == 200
    assert payment_import.json()["success"] == 1

    payment_update_file = build_excel(
        payment_headers,
        [[contract_code, 1, "第一期-更新", (date.today() + timedelta(days=7)).isoformat(), "5000", date.today().isoformat(), "3000", "已提报", "更新首付款", "更新导入"]],
    )
    payment_update = client.post(
        "/api/import/excel/payments",
        params={"duplicate_action": "update"},
        files={
            "file": (
                "payments-update.xlsx",
                payment_update_file,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert payment_update.status_code == 200
    assert payment_update.json()["success"] == 1

    payment_lookup = client.get("/api/payments")
    imported_payment = next(item for item in payment_lookup.json() if item["contract_id"] == imported_contract["id"] and item["seq"] == 1)
    assert imported_payment["phase"] == "第一期-更新"
    assert imported_payment["payment_status"] == "已提报"
    assert as_decimal(imported_payment["pending_amount"]) == Decimal("2000")


def test_ai_screenshot_parse_and_confirm_import(client: TestClient, monkeypatch: pytest.MonkeyPatch) -> None:
    from backend.app.routers import imports as imports_router

    parsed_data = {
        "contract": {
            "contract_code": unique_code("AI-HT"),
            "contract_name": "AI识别合同",
            "vendor": "AI供应商",
            "amount": "24000",
            "status": "签订",
            "project_code": unique_code("AI-PJ"),
            "project_name": "AI导入项目",
            "payment_direction": "支出",
        },
        "items": [
            {"seq": 1, "item_name": "AI标的", "quantity": "2", "unit": "项", "unit_price": "8000", "amount": "16000"},
            {"seq": 2, "item_name": "AI运维", "quantity": "1", "unit": "项", "unit_price": "8000", "amount": "8000"},
        ],
        "payment_plans": [
            {"seq": 1, "phase": "首付", "planned_date": (date.today() + timedelta(days=5)).isoformat(), "planned_amount": "12000", "payment_status": "未付"},
            {"seq": 2, "phase": "尾款", "planned_date": (date.today() + timedelta(days=30)).isoformat(), "planned_amount": "12000", "payment_status": "未付"},
        ],
        "changes": [
            {"seq": 1, "change_date": date.today().isoformat(), "change_info": "补充服务", "before_content": "基础服务", "after_content": "基础服务+运维", "change_description": "AI识别出的变更"},
        ],
    }

    monkeypatch.setattr(imports_router, "parse_screenshots", lambda _images: (parsed_data, []))

    parse_response = client.post(
        "/api/import/screenshot",
        files=[("files", ("screenshot.png", b"fake-image", "image/png"))],
    )
    assert parse_response.status_code == 200
    assert parse_response.json()["parsed_data"]["contract"]["contract_name"] == "AI识别合同"
    assert parse_response.json()["uncertain_fields"] == []

    confirm_response = client.post("/api/import/screenshot/confirm", json={"parsed_data": parsed_data})
    assert confirm_response.status_code == 200
    confirm_data = confirm_response.json()
    assert confirm_data["message"] == "导入成功"

    contract_detail = client.get(f"/api/contracts/{confirm_data['contract_id']}")
    assert contract_detail.status_code == 200
    contract_data = contract_detail.json()
    assert contract_data["contract_name"] == "AI识别合同"
    assert len(contract_data["items"]) == 2
    assert len(contract_data["payments"]) == 2
    assert len(contract_data["changes"]) == 1
